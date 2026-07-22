from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from sqlalchemy import select
import io

from app.database import get_db
from app.models.user import User
from app.models.material import BOMItem, Material
from app.schemas.material import (
    MaterialCategoryCreate,
    MaterialCategoryResponse,
    MaterialCreate,
    MaterialResponse,
    BOMItemCreate,
    BOMItemResponse,
    BOMSummaryResponse,
    BOMGenerateResponse,
)
from app.auth import get_current_user
from app.rbac import verify_project_access
from app.models.project import Project
from app.services import material_service
from app.ws import ws_manager

router = APIRouter(prefix="/materials", tags=["物料"])


@router.get("/categories", response_model=list[MaterialCategoryResponse])
async def list_categories(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """公开端点：物料分类列表（平台公共建材目录）"""
    categories = await material_service.get_categories(db)
    return [MaterialCategoryResponse.model_validate(c) for c in categories]


@router.post("/categories", response_model=MaterialCategoryResponse, status_code=status.HTTP_201_CREATED)
async def create_category(
    data: MaterialCategoryCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    category = await material_service.create_category(db, data.model_dump())
    return MaterialCategoryResponse.model_validate(category)


@router.get("", response_model=list[MaterialResponse])
async def list_materials(
    current_user: User = Depends(get_current_user),
    category_id: str | None = Query(None),
    keyword: str | None = Query(None, description="按名称/SKU/品牌模糊搜索"),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
):
    if keyword:
        materials = await material_service.search_materials(db, keyword, skip, limit)
    else:
        materials = await material_service.get_materials(db, category_id, skip, limit)
    return [MaterialResponse.model_validate(m) for m in materials]


@router.get("/{material_id}", response_model=MaterialResponse)
async def get_material(
    material_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """公开端点：物料详情（平台公共建材目录）"""
    material = await material_service.get_material_by_id(db, material_id)
    if not material:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="物料不存在")
    return MaterialResponse.model_validate(material)


@router.post("", response_model=MaterialResponse, status_code=status.HTTP_201_CREATED)
async def create_material(
    data: MaterialCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    material = await material_service.create_material(db, data.model_dump())
    return MaterialResponse.model_validate(material)


@router.post("/bom", response_model=BOMItemResponse, status_code=status.HTTP_201_CREATED)
async def add_bom_item(
    data: BOMItemCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await verify_project_access(project_id=data.project_id, current_user=current_user, db=db)
    bom_item = await material_service.add_bom_item(db, data.model_dump())
    from sqlalchemy import select as sa_select

    result = await db.execute(
        sa_select(BOMItem)
        .where(BOMItem.id == bom_item.id)
        .options(selectinload(BOMItem.material).selectinload(Material.category))
    )
    bom_item = result.scalar_one()
    resp = BOMItemResponse.model_validate(bom_item)
    await ws_manager.broadcast_to_project(bom_item.project_id, "bom.item_added", resp.model_dump())
    return resp


@router.post("/bom/generate/{project_id}", response_model=BOMGenerateResponse, status_code=status.HTTP_201_CREATED)
async def generate_bom(
    project_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """F6 BOM 自动生成 — 基于项目房间面积/类型按标准用量生成"""
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目不存在")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权访问该项目")
    try:
        items = await material_service.generate_bom_for_project(db, project_id)
    except ValueError as e:
        if str(e) == "PROJECT_ALREADY_HAS_BOM":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="该项目已有 BOM 物料清单，请先清空再生成",
            )
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))

    if not items:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="项目下未找到房间数据，无法生成 BOM",
        )

    total_price = round(sum(item.total_price for item in items), 2)
    resp = BOMGenerateResponse(
        project_id=project_id,
        generated_count=len(items),
        total_price=total_price,
        items=[BOMItemResponse.model_validate(item) for item in items],
    )
    await ws_manager.broadcast_to_project(
        project_id, "bom.generated", resp.model_dump()
    )
    return resp


@router.get("/bom/{project_id}", response_model=list[BOMItemResponse])
async def get_project_bom(
    project_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目不存在")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权访问该项目")
    bom_items = await material_service.get_project_bom(db, project_id)
    return [BOMItemResponse.model_validate(item) for item in bom_items]


@router.get("/bom/{project_id}/summary", response_model=BOMSummaryResponse)
async def get_bom_summary(
    project_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """BOM 汇总（按品类聚合）"""
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目不存在")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权访问该项目")
    summary = await material_service.get_bom_summary(db, project_id)
    if not summary:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="该项目暂无 BOM 数据",
        )
    return BOMSummaryResponse(**summary)


@router.get("/bom/{project_id}/export")
async def export_bom_excel(
    project_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """F7 BOM Excel 导出"""
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目不存在")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权访问该项目")
    bom_items = await material_service.get_project_bom(db, project_id)
    if not bom_items:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="该项目暂无 BOM 数据，无法导出",
        )
    # v1.1.14: 延迟导入 openpyxl，减少应用启动时间和内存占用
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM 物料清单"

    headers = ["序号", "物料名称", "SKU", "品牌", "规格", "分类", "数量", "单位", "单价(元)", "总价(元)", "备注"]
    ws.append(headers)

    for i, item in enumerate(bom_items, 1):
        mat = item.material
        ws.append([
            i,
            mat.name if mat else "-",
            mat.sku if mat else "-",
            mat.brand if mat else "-",
            mat.spec if mat else "-",
            mat.category.name if mat and mat.category else "-",
            item.quantity,
            mat.unit if mat else "-",
            item.unit_price,
            item.total_price,
            item.note or "",
        ])

    total_row = len(bom_items) + 2
    ws.cell(row=total_row, column=9, value="合计：")
    ws.cell(row=total_row, column=10, value=sum(item.total_price for item in bom_items))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=bom-{project_id[:8]}.xlsx"},
    )


@router.delete("/bom/{bom_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_bom_item(
    bom_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import select as sa_select

    result = await db.execute(sa_select(BOMItem).where(BOMItem.id == bom_id))
    bom_item = result.scalar_one_or_none()
    if not bom_item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="BOM项不存在")
    project_id = bom_item.project_id
    await verify_project_access(project_id=project_id, current_user=current_user, db=db)
    deleted = await material_service.delete_bom_item(db, bom_id)
    if not deleted:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="BOM项不存在")
    await ws_manager.broadcast_to_project(project_id, "bom.item_deleted", {"id": bom_id})


@router.post("/bom/auto-match")
async def auto_match_bom(
    project_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """智能 BOM 物料匹配：根据项目 BOM 中的物料名称/规格，自动匹配物料库中的标准物料

    匹配策略（3 级降级）：
    1. SKU 精确匹配：BOM 中的 sku 与物料库 sku 完全一致
    2. 名称模糊匹配：BOM 中物料名称包含物料库名称或反之（双向）
    3. 品类 + 规格匹配：品类相同且规格相似（单位一致）

    返回每个 BOM 项的匹配结果（match_level: exact/fuzzy/category/unmatched + matched_material + confidence）
    """
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目不存在")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权访问该项目")
    from sqlalchemy import select as sa_select

    # 获取项目的 BOM（预加载 material + category 避免 N+1 懒加载）
    bom_result = await db.execute(
        sa_select(BOMItem)
        .options(selectinload(BOMItem.material).selectinload(Material.category))
        .where(BOMItem.project_id == project_id)
    )
    bom_items = bom_result.scalars().all()

    if not bom_items:
        return {
            "project_id": project_id,
            "matched": 0,
            "total": 0,
            "results": [],
            "reply": "该项目暂无 BOM 数据",
        }

    # 加载物料库（预加载 category 避免懒加载）
    mat_result = await db.execute(
        sa_select(Material).options(selectinload(Material.category))
    )
    all_materials = mat_result.scalars().all()

    # 一次性建立三类索引（O(M)），替代每个 BOM 项遍历全表
    materials_by_sku: dict[str, Material] = {}
    materials_by_name: list[Material] = []  # 模糊匹配需遍历，保留列表
    materials_by_cat_unit: dict[tuple[str, str], list[Material]] = {}

    for m in all_materials:
        if m.sku:
            sku = m.sku.strip()
            if sku:
                materials_by_sku[sku] = m
        if m.name:
            materials_by_name.append(m)
        m_cat = (m.category.name if m.category else "").strip()
        m_unit = (m.unit or "").strip()
        materials_by_cat_unit.setdefault((m_cat, m_unit), []).append(m)

    results = []
    matched = 0

    for item in bom_items:
        mat_name = (
            item.material.name if item.material and item.material.name else ""
        ).strip()
        mat_sku = (
            item.material.sku if item.material and item.material.sku else ""
        ).strip()
        mat_category = (
            item.material.category.name
            if item.material and item.material.category and item.material.category.name
            else ""
        ).strip()
        mat_unit = (
            item.material.unit if item.material and item.material.unit else ""
        ).strip()

        match = None
        match_level = "unmatched"
        confidence = 0.0

        # Level 1: SKU 精确匹配 — O(1) dict 查找
        if mat_sku and mat_sku in materials_by_sku:
            match = materials_by_sku[mat_sku]
            match_level = "exact"
            confidence = 1.0

        # Level 2: 名称模糊匹配 — O(M) 遍历但可短路
        if not match and mat_name:
            for m in materials_by_name:
                if mat_name in m.name or m.name in mat_name:
                    match = m
                    match_level = "fuzzy"
                    confidence = 0.7
                    break

        # Level 3: 品类 + 规格匹配 — O(1) dict 查找取首个
        if not match and mat_category:
            candidates = materials_by_cat_unit.get((mat_category, mat_unit))
            if candidates:
                match = candidates[0]
                match_level = "category"
                confidence = 0.4

        if match:
            matched += 1

        results.append(
            {
                "bom_item_id": item.id,
                "material_name": mat_name,
                "material_sku": mat_sku,
                "category": mat_category,
                "quantity": item.quantity,
                "match_level": match_level,
                "confidence": confidence,
                "matched_material": {
                    "id": match.id if match else None,
                    "name": match.name if match else None,
                    "sku": match.sku if match else None,
                    "unit_price": float(match.unit_price)
                    if match and match.unit_price
                    else None,
                    "unit": match.unit if match else None,
                }
                if match
                else None,
            }
        )

    return {
        "project_id": project_id,
        "matched": matched,
        "total": len(results),
        "match_rate": round(matched / max(len(results), 1) * 100, 2),
        "results": results,
        "reply": f"BOM 自动匹配完成：{matched}/{len(results)} 项匹配成功（匹配率 {round(matched / max(len(results), 1) * 100, 1)}%）",
    }
