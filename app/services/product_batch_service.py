"""批量产品服务 — Excel/CSV 解析 + 校验 + 批量写入"""

import csv
import io
import json
import logging
import uuid

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.product import Product
from app.models.user import User
from app.models.procurement import Supplier
from app.schemas.product_batch import (
    BatchProductRow,
    BatchProductResult,
    BatchUploadResponse,
)

logger = logging.getLogger("ihome")

# 允许的分类代码
VALID_CATEGORIES = {
    "tile", "flooring", "cabinet", "paint", "lighting",
    "appliance", "curtain", "custom_furniture", "service", "other",
}

# 分类中文名 → 代码映射（用于解析模板中的中文分类名）
CATEGORY_CN_TO_CODE = {
    "瓷砖": "tile", "地板": "flooring", "橱柜": "cabinet",
    "涂料": "paint", "灯具": "lighting", "家电": "appliance",
    "窗帘": "curtain", "定制家具": "custom_furniture", "服务": "service",
    "其他": "other",
}

VALID_STOCK_STATUS = {"in_stock", "pre_order", "out_of_stock"}


async def parse_excel_file(file_content: bytes) -> list[BatchProductRow]:
    """解析 Excel 文件（.xlsx），返回产品行列表"""
    workbook = load_workbook(filename=io.BytesIO(file_content), read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Excel 文件至少需要包含表头和一行数据")

    # 解析表头 -> 列索引映射
    header = [str(h).strip() if h else "" for h in rows[0]]
    col_map = {name: idx for idx, name in enumerate(header)}

    products: list[BatchProductRow] = []
    for row_idx, row in enumerate(rows[1:], start=1):
        try:
            product = _parse_row_to_product(row, col_map, row_idx)
            if product:
                products.append(product)
        except Exception as e:
            logger.warning(f"Excel 第 {row_idx + 1} 行解析失败: {e}")

    workbook.close()
    return products


async def parse_csv_file(file_content: bytes) -> list[BatchProductRow]:
    """解析 CSV 文件，返回产品行列表"""
    text = file_content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    products: list[BatchProductRow] = []

    for row_idx, row in enumerate(reader, start=1):
        # 构建列映射（CSV 用列名直接访问）
        col_map = {k.strip(): i for i, k in enumerate(row.keys())}
        values = list(row.values())
        try:
            product = _parse_row_to_product(values, col_map, row_idx)
            if product:
                products.append(product)
        except Exception as e:
            logger.warning(f"CSV 第 {row_idx + 1} 行解析失败: {e}")

    return products


def _parse_row_to_product(row: tuple, col_map: dict, row_idx: int) -> BatchProductRow | None:
    """将一行数据解析为 BatchProductRow，返回 None 表示跳过空行"""

    def _get(col_names: list[str], default=""):
        for cn in col_names:
            if cn in col_map:
                val = row[col_map[cn]]
                return str(val).strip() if val else default
        return default

    name = _get(["名称", "产品名称", "name", "product_name"])
    if not name:
        return None  # 跳过空行

    category_raw = _get(["分类", "类别", "category"], "other")
    category = CATEGORY_CN_TO_CODE.get(category_raw, category_raw)
    if category not in VALID_CATEGORIES:
        category = "other"

    # 价格
    price_min = None
    price_max = None
    price_raw = _get(["最低价", "最低价格", "price_min", "价格", "price"])
    price_max_raw = _get(["最高价", "最高价格", "price_max"])
    if price_raw:
        try:
            price_min = float(price_raw)
        except ValueError:
            pass
    if price_max_raw:
        try:
            price_max = float(price_max_raw)
        except ValueError:
            pass

    unit = _get(["单位", "unit"], "个")
    description = _get(["描述", "产品描述", "description"])
    tags = _get(["标签", "tags"])
    stock_status = _get(["库存状态", "stock_status"], "in_stock")
    if stock_status not in VALID_STOCK_STATUS:
        stock_status = "in_stock"
    cover_image = _get(["图片", "图片URL", "cover_image", "image_url"])

    return BatchProductRow(
        name=name,
        category=category,
        description=description or None,
        price_min=price_min,
        price_max=price_max,
        unit=unit or "个",
        tags=tags or None,
        stock_status=stock_status,
        cover_image=cover_image or None,
    )


async def batch_create_products(
    db: AsyncSession,
    current_user: User,
    supplier: Supplier | None,
    products: list[BatchProductRow],
) -> BatchUploadResponse:
    """批量创建产品，返回每个产品的创建结果"""
    results: list[BatchProductResult] = []
    success_count = 0
    failed_count = 0

    for idx, row in enumerate(products):
        row_num = idx + 1
        try:
            # tags 字符串 → 列表
            tags_list = None
            if row.tags:
                tags_list = [t.strip() for t in row.tags.replace("，", ",").split(",") if t.strip()]

            product = Product(
                id=str(uuid.uuid4()),
                user_id=current_user.id,
                supplier_id=supplier.id if supplier else "",
                name=row.name,
                category=row.category,
                description=row.description,
                price_min=row.price_min,
                price_max=row.price_max,
                unit=row.unit,
                cover_image=row.cover_image,
                images=None,
                tags=json.dumps(tags_list, ensure_ascii=False) if tags_list else None,
                specs=None,
                stock_status=row.stock_status,
                status="draft",
                ai_assisted=False,  # 先不标记，由后台任务单独处理
            )
            db.add(product)
            await db.flush()

            results.append(BatchProductResult(
                row=row_num,
                name=row.name,
                success=True,
                product_id=product.id,
            ))
            success_count += 1
        except Exception as e:
            results.append(BatchProductResult(
                row=row_num,
                name=row.name,
                success=False,
                error=str(e)[:200],
            ))
            failed_count += 1

    await db.commit()

    batch_id = str(uuid.uuid4())
    return BatchUploadResponse(
        total=len(products),
        success_count=success_count,
        failed_count=failed_count,
        results=results,
        batch_id=batch_id,
        ai_jobs_pending=success_count,
    )


def generate_template_excel() -> bytes:
    """生成示例 Excel 模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "产品批量导入模板"

    # 样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 表头
    headers = [
        "名称", "分类", "最低价", "最高价", "单位",
        "描述", "标签", "库存状态", "图片URL",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 列宽
    col_widths = [30, 12, 10, 10, 8, 40, 25, 12, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # 示例数据
    sample_data = [
        ["800×800 灰色防滑地砖", "瓷砖", 58, 68, "㎡", "佛山产 R10防滑 亮面",
         "防滑,灰色,客厅,地砖", "in_stock", "https://example.com/img/floor.jpg"],
        ["北欧简约三人沙发", "定制家具", 4500, 5500, "件", "棉麻面料 实木框架 2100×850×800mm", "北欧,简约,沙发,客厅", "in_stock", ""],
        ["LED 无主灯套装", "灯具", 1680, 2680, "套", "全屋 12灯 4000K 含轨道灯+筒灯+射灯", "LED,无主灯,智能", "in_stock", ""],
        ["净味乳胶漆 18L", "涂料", 580, 680, "桶", "净味配方 即刷即住", "乳胶漆,净味", "in_stock", ""],
        ["定制橱柜 地柜+吊柜", "橱柜", 1880, 2280, "m", "E0级颗粒板 石英石台面", "橱柜,定制,厨房", "in_stock", ""],
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # 说明 sheet
    ws2 = wb.create_sheet("填写说明")
    instructions = [
        ["字段", "说明", "可选值/示例", "是否必填"],
        ["名称", "产品名称，含规格信息", "800×800 灰色防滑地砖", "是"],
        ["分类", "产品分类中文名", "瓷砖/地板/橱柜/涂料/灯具/家电/窗帘/定制家具/服务/其他", "是"],
        ["最低价", "最低销售价格（数字）", "58", "否"],
        ["最高价", "最高销售价格（数字）", "68", "否"],
        ["单位", "计价单位", "㎡/m/个/台/件/套/桶/次", "否，默认「个」"],
        ["描述", "产品描述、卖点、材质等", "佛山产 R10防滑 亮面", "否"],
        ["标签", "逗号分隔的标签", "防滑,灰色,客厅", "否"],
        ["库存状态", "库存状态", "in_stock(有货)/pre_order(预售)/out_of_stock(售罄)", "否，默认 in_stock"],
        ["图片URL", "产品主图链接", "https://example.com/img.jpg", "否"],
    ]
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 40
    ws2.column_dimensions["D"].width = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
